import requests
from io import BytesIO
import openpyxl
import json

url = "https://docs.google.com/spreadsheets/d/1qRdicMJkMXNFBKCoeJosPCbsMa7scVxr/export?format=xlsx"
print("Downloading traces...")
r = requests.get(url)
print(f"Status: {r.status_code}, Size: {len(r.content)} bytes")

wb = openpyxl.load_workbook(BytesIO(r.content))
print(f"Sheets: {wb.sheetnames}")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*60}")
    print(f"Sheet: {sheet_name}")
    print(f"Dimensions: {ws.dimensions}, Rows: {ws.max_row}, Cols: {ws.max_column}")
    print(f"{'='*60}")
    
    # Print all rows
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        print(f"Row {i}: {row}")
        if i > 30:  # Safety limit
            print("... (truncated)")
            break

# Save to JSON for later use
data = {}
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {h: v for h, v in zip(headers, row) if h is not None}
        if any(v is not None for v in row_dict.values()):
            rows.append(row_dict)
    data[sheet_name] = {"headers": headers, "rows": rows}

with open("data/traces.json", "w", encoding="utf-8") as f:
    json.dump(data, f, indent=2, default=str)
print("\nSaved to data/traces.json")
